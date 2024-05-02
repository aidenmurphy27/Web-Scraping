import openpyxl as xl
from openpyxl.styles import Font

#create a new excel document
wb = xl.Workbook()

ws = wb.active


ws.title = 'First Sheet'

ws2 = wb.create_sheet(index=1,title='Second Sheet')

#write content to a cell
ws['A1'] = 'Invoice'

ws['A1'].font = Font(name='Times New Roman', size=24,italic=False,bold=True)

headerFont = Font(name='Times New Roman', size=24,italic=False,bold=True)

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Allignment'

ws.merge_cells('A1:B1')

#ws['A2'].font = headerFont
#ws['A3'].font = headerFont
#ws['A4'].font = headerFont

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = Font(size=16,bold=True)

ws['B8'] = '=SUM(B2:B7)'






wb.save("PythontoExcel.xlsx")

#Class Example

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

ws2 = wb.active

ws2['A44'] = 'Total'
ws2['A44'].font = Font(size=16,bold=True)
